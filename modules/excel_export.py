from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils import get_column_letter


def export_matrix_to_excel(matrix_records: List[Dict[str, Any]], template_path: Path | None = None) -> bytes:
    """Export matrix to Excel. If template_path exists, we try to write starting at the first empty row.

    V1.1: dùng layout đơn giản: header ở hàng 1 với đúng keys hiện có trong matrix_records.
    """
    if template_path and template_path.exists():
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        # find starting row: first empty row after header
        start_row = ws.max_row + 1
        # Determine header row: assume row 1
        headers = [cell.value for cell in ws[1]]
        headers = [h if h is not None else "" for h in headers]

        # If template has no headers, fallback
        if not any(str(h).strip() for h in headers):
            headers = list(matrix_records[0].keys()) if matrix_records else []
            for j, h in enumerate(headers, start=1):
                ws.cell(row=1, column=j, value=h)
            start_row = 2

        # write
        for i, rec in enumerate(matrix_records):
            r = start_row + i
            for j, h in enumerate(headers, start=1):
                key = str(h).strip()
                if not key:
                    continue
                ws.cell(row=r, column=j, value=rec.get(key, ""))

    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(matrix_records[0].keys()) if matrix_records else []
        for j, h in enumerate(headers, start=1):
            ws.cell(row=1, column=j, value=h)
        for i, rec in enumerate(matrix_records):
            r = 2 + i
            for j, h in enumerate(headers, start=1):
                ws.cell(row=r, column=j, value=rec.get(h, ""))

        # basic widths
        for j in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(j)].width = min(40, max(12, len(str(headers[j - 1])) + 2))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
